from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from pipeline.ids import content_hash, stable_id
from pipeline.metrics import (
    academic_accuracy,
    difficulty_band,
    ineffective_distractors,
    nps,
    evaluation_rate,
    participation_rate,
    point_biserial,
    public_value,
    reinforce,
    rolling_average,
    wilson_interval,
)
from pipeline.models import Question, QuestionKind, Response, Session
from pipeline.parser import UnknownSchemaError, parse_workbook


def test_stable_ids_and_hashes_are_order_independent() -> None:
    assert stable_id("session", " deck ", "27/05/2026") == stable_id(
        "session", "deck", "27/05/2026"
    )
    assert content_hash({"b": 2, "a": 1}) == content_hash({"a": 1, "b": 2})


def test_reference_metrics_27_may_2026() -> None:
    assert participation_rate(208, 245) == pytest.approx(0.849, abs=0.0005)
    assert participation_rate(9, 35) == pytest.approx(0.257, abs=0.0005)
    assert academic_accuracy(142, 166) == pytest.approx(0.855, abs=0.0005)


def test_synthetic_reference_fixture_reproduces_published_metrics(synthetic_reference_xlsx: Path) -> None:
    questions, responses = parse_workbook(synthetic_reference_xlsx, "27-05-2026")
    kinds = {q.question_id: q.kind for q in questions}
    academic = [r for r in responses if kinds[r.question_id] == QuestionKind.ACADEMIC]
    evaluations = [r for r in responses if kinds[r.question_id] == QuestionKind.EVALUATION]
    participants = len({r.participant_id for r in responses})
    assert participation_rate(len(responses), participants, 7) == pytest.approx(0.849, abs=0.0005)
    assert evaluation_rate(len({r.participant_id for r in evaluations}), participants) == pytest.approx(0.257, abs=0.0005)
    assert academic_accuracy(sum(r.is_correct is True for r in academic), len(academic)) == pytest.approx(0.855, abs=0.0005)


def test_metric_policy_and_statistics() -> None:
    assert difficulty_band(0.29) == "very_hard"
    assert difficulty_band(0.30) == "hard"
    assert difficulty_band(0.50) == "medium"
    assert difficulty_band(0.70) == "easy"
    assert difficulty_band(0.85) == "very_easy"
    assert rolling_average(list(range(1, 11)), window=8)[-1] == pytest.approx(6.5)
    low, high = wilson_interval(142, 166)
    assert 0.79 < low < 0.81 and 0.90 < high < 0.91
    assert point_biserial([1] * 10 + [0] * 9, list(range(19))) is None
    assert point_biserial([1] * 10 + [0] * 10, list(range(20))) is not None
    assert ineffective_distractors({"A": 90, "B": 4, "C": 6}, "A") == ["B"]
    assert nps([10, 9, 8, 7, 6, 3]) == pytest.approx(0.0)
    assert reinforce(0.59, 30) is True
    assert reinforce(0.59, 29) is False
    assert public_value(4.5, 4) is None
    assert public_value(4.5, 5) == 4.5


def _make_portuguese_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    ws.append(["ID do participante", "Pergunta", "Resposta", "Correta", "Tipo"])
    ws.append(["u1", "Participante", "Aluno", "", "perfil"])
    ws.append(["u1", "Pergunta 1", "A", "Sim", "quiz"])
    ws.append(["u2", "Pergunta 1", "B", "Não", "quiz"])
    ws.append(["u1", "Avaliação geral", 9, "", "avaliação"])
    wb.save(path)


def test_parser_recognizes_portuguese_headers_and_excludes_non_academic(tmp_path: Path) -> None:
    path = tmp_path / "respostas.xlsx"
    _make_portuguese_workbook(path)
    questions, responses = parse_workbook(path, "deck-1")
    assert len(responses) == 4
    kinds = {question.title: question.kind for question in questions}
    assert kinds["Participante"] == QuestionKind.PROFILE
    assert kinds["Avaliação geral"] == QuestionKind.EVALUATION
    academic = [r for r in responses if r.is_correct is not None]
    assert academic_accuracy(sum(r.is_correct is True for r in academic), len(academic)) == 0.5


def test_parser_recognizes_english_headers_and_fails_closed(tmp_path: Path) -> None:
    good = tmp_path / "answers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Participant ID", "Question", "Answer", "Correct", "Question type"])
    ws.append(["x", "Q", "B", "true", "quiz"])
    wb.save(good)
    assert parse_workbook(good, "deck")[1][0].is_correct is True

    bad = tmp_path / "unknown.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Foo", "Bar", "Baz"])
    ws.append([1, 2, 3])
    wb.save(bad)
    with pytest.raises(UnknownSchemaError):
        parse_workbook(bad, "deck")


def test_participants_are_distinct_per_session() -> None:
    sessions = [
        Session("s1", "p", date(2026, 5, 27), 10, 5, complete=True),
        Session("s2", "p", date(2026, 6, 3), 12, 5, complete=True),
    ]
    assert sum(s.participants for s in sessions) == 22
