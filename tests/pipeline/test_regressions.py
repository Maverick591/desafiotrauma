from __future__ import annotations

import json
from zipfile import ZipFile
from datetime import date, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from pipeline.ai import AIClassifier, Classification, minimize_for_ai, redact_pii
from pipeline.mentimeter import PresentationRef, select_presentations
from pipeline.models import Presentation, Question, QuestionKind, Response, Session
from pipeline.orchestrator import (
    Pipeline,
    build_public_snapshot,
    published_corpus,
    responses_with_answer_keys,
    session_date_from_responses,
    validate_public_snapshot,
)
from pipeline.parser import InvalidWorkbookError, UnknownSchemaError, parse_workbook
from pipeline.persistence import LocalRepository, ManualImport, SupabaseRepository, question_payload
from pipeline.privacy import valid_response_count
from pipeline.reports import write_report


def test_incremental_presentation_filter_and_empty_discovery_fail_closed(tmp_path: Path) -> None:
    refs = [PresentationRef("a", "Desafio Trauma - 01/01/2026", "/a")]
    assert select_presentations(refs, "incremental", set(), manual_id="a") == refs

    class EmptyClient:
        def discover(self):
            return []

    class Repo:
        def states(self): return {}

    with pytest.raises(RuntimeError, match="No presentations discovered"):
        Pipeline(client=EmptyClient(), repository=Repo(), workdir=tmp_path).sync("incremental", dry_run=True)


def test_snapshot_v1_contract_units_privacy_and_distinct_participants() -> None:
    presentation = Presentation("p", "Desafio Trauma - 01/01/2026", date(2026, 1, 1), "/p")
    sessions = [
        Session("small", "p", date(2026, 1, 1), 4, 1, True),
        Session("eligible", "p", date(2026, 1, 8), 5, 1, True),
    ]
    questions = [Question("q", "p", 1, "Q", QuestionKind.ACADEMIC, ("A", "B"), (0,), topic="Choque")]
    responses = [
        Response(f"r{i}", "eligible", "q", f"user-{i}", "A", i < 4) for i in range(5)
    ] + [Response("duplicate", "eligible", "q", "user-0", "A", True)]
    snapshot = build_public_snapshot([presentation], sessions, questions, responses, public_files=[{"name": "X", "url": "/x"}])
    assert set(snapshot) == {"metadata", "filters", "overview", "participation", "learning", "experience", "topics", "metric_dictionary", "public_files"}
    assert snapshot["overview"]["accuracy_rate"] == pytest.approx(83.3, abs=0.1)
    assert snapshot["participation"]["total_participants"] == 5
    assert snapshot["participation"]["trend"] == [{"label": "08/01/2026", "participants": 5, "responses": 5}]
    assert len(snapshot["overview"]["trend"]) == 1  # n<5 session excluded from moving series
    validate_public_snapshot(snapshot)


def test_public_totals_and_profile_denominator_cannot_reveal_suppressed_groups(tmp_path: Path) -> None:
    presentation = Presentation("p", "Desafio Trauma - 01/01/2026", date(2026, 1, 1), "/p")
    sessions = [
        Session("small", "p", date(2026, 1, 1), 4, 2, True),
        Session("visible", "p", date(2026, 1, 8), 5, 2, True),
    ]
    academic = Question("qa", "p", 1, "Q", QuestionKind.ACADEMIC)
    profile = Question("qp", "p", 2, "Perfil", QuestionKind.PROFILE)
    responses = []
    for session_id, count, profile_value in (("small", 4, "Oculto"), ("visible", 5, "Cirurgião")):
        for index in range(count):
            person = f"{session_id}-{index}"
            responses.extend([
                Response(f"{session_id}-a-{index}", session_id, "qa", person, "A", True),
                Response(f"{session_id}-p-{index}", session_id, "qp", person, profile_value, None),
            ])

    snapshot = build_public_snapshot([presentation], sessions, [academic, profile], responses)
    assert snapshot["overview"]["participants"] == 5
    assert snapshot["overview"]["responses"] == 10
    assert snapshot["participation"]["by_profile"] == [{"label": "Cirurgião", "value": 100.0}]
    assert all(row["label"] != "01/01/2026" for row in snapshot["participation"]["trend"])

    report = write_report(tmp_path / "public.xlsx", [presentation], sessions, [academic, profile], responses, public=True)
    workbook = load_workbook(report, data_only=True)
    summary = dict(workbook["Resumo"].iter_rows(min_row=2, values_only=True))
    assert summary["Participantes (soma por sessão)"] == 5
    assert workbook["Encontros"].max_row == 2
    assert workbook["Encontros"]["C2"].value == 5


def test_profile_views_are_disabled_when_any_complement_is_below_k() -> None:
    presentation = Presentation("p", "Desafio Trauma - 01/01/2026", date(2026, 1, 1), "/p")
    session = Session("s", "p", date(2026, 1, 1), 9, 2, True)
    academic = Question("qa", "p", 1, "Q", QuestionKind.ACADEMIC)
    profile = Question("qp", "p", 2, "Perfil", QuestionKind.PROFILE)
    responses = []
    for index in range(9):
        person = f"u-{index}"
        responses.extend([
            Response(f"a-{index}", "s", "qa", person, "A", True),
            Response(f"p-{index}", "s", "qp", person, "Perfil A" if index < 5 else "Perfil B", None),
        ])
    snapshot = build_public_snapshot([presentation], [session], [academic, profile], responses)
    assert snapshot["overview"]["participants"] == 9
    assert snapshot["participation"]["by_profile"] == []
    assert snapshot["filters"]["profiles"] == ["Todos os perfis"]
    assert not any(view["filters"].get("profile") for view in snapshot["filters"]["views"])


def test_question_below_k_is_excluded_from_public_details_and_totals(tmp_path: Path) -> None:
    presentation = Presentation("p", "Desafio Trauma - 01/01/2026", date(2026, 1, 1), "/p")
    session = Session("s", "p", date(2026, 1, 1), 5, 2, True)
    visible = Question("visible", "p", 1, "Visível", QuestionKind.ACADEMIC)
    hidden = Question("hidden", "p", 2, "Oculta", QuestionKind.ACADEMIC)
    responses = [Response(f"v-{i}", "s", "visible", f"u-{i}", "A", True) for i in range(5)]
    responses += [Response(f"h-{i}", "s", "hidden", f"u-{i}", "B", False) for i in range(4)]

    snapshot = build_public_snapshot([presentation], [session], [visible, hidden], responses)
    assert snapshot["learning"]["questions"] == 1
    assert snapshot["learning"]["answers"] == 5
    assert snapshot["overview"]["responses"] == 5
    assert [row["question"] for row in snapshot["learning"]["question_performance"]] == ["Visível"]

    report = write_report(tmp_path / "public.xlsx", [presentation], [session], [visible, hidden], responses, public=True)
    workbook = load_workbook(report, data_only=True)
    assert workbook["Questões agregadas"].max_row == 2
    summary = dict(workbook["Resumo"].iter_rows(min_row=2, values_only=True))
    assert summary["Participação"] == 1


def test_people_exclusive_to_hidden_question_are_removed_from_public_total() -> None:
    presentation = Presentation("p", "Desafio Trauma - 01/01/2026", date(2026, 1, 1), "/p")
    session = Session("s", "p", date(2026, 1, 1), 9, 2, True)
    visible = Question("visible", "p", 1, "Visível", QuestionKind.ACADEMIC)
    hidden = Question("hidden", "p", 2, "Oculta", QuestionKind.ACADEMIC)
    responses = [Response(f"v-{i}", "s", "visible", f"visible-{i}", "A", True) for i in range(5)]
    responses += [Response(f"h-{i}", "s", "hidden", f"hidden-{i}", "B", False) for i in range(4)]
    snapshot = build_public_snapshot([presentation], [session], [visible, hidden], responses)
    assert snapshot["overview"]["participants"] == 5
    assert snapshot["overview"]["responses"] == 5


def test_snapshot_contract_accepts_legitimate_one_percent() -> None:
    presentation = Presentation("p", "Desafio Trauma - 01/01/2026", date(2026, 1, 1), "/p")
    sessions = [Session("s", "p", date(2026, 1, 1), 100, 1, True)]
    questions = [Question("q", "p", 1, "Q", QuestionKind.ACADEMIC)]
    responses = [Response(f"r{i}", "s", "q", f"u{i}", "A", i == 0) for i in range(100)]
    snapshot = build_public_snapshot([presentation], sessions, questions, responses)
    assert snapshot["learning"]["accuracy_rate"] == 1.0
    validate_public_snapshot(snapshot)


def test_parser_requires_real_participant_and_ids_ignore_row_position(tmp_path: Path) -> None:
    missing = tmp_path / "missing-participant.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Question", "Answer", "Correct"]); ws.append(["Q", "A", "true"]); wb.save(missing)
    with pytest.raises(UnknownSchemaError):
        parse_workbook(missing, "p")

    def make(path: Path, blank_first: bool):
        wb = Workbook(); ws = wb.active
        ws.append(["Participant ID", "Slide", "Question", "Answer", "Correct", "Question type"])
        if blank_first: ws.append([None] * 6)
        ws.append(["u", 2, "Mesmo título", "A", "true", "quiz"])
        wb.save(path)

    first, second = tmp_path / "first.xlsx", tmp_path / "second.xlsx"
    make(first, False); make(second, True)
    q1, r1 = parse_workbook(first, "p"); q2, r2 = parse_workbook(second, "p")
    assert q1[0].question_id == q2[0].question_id
    assert r1[0].response_id == r2[0].response_id

    duplicate = tmp_path / "duplicate.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Participant ID", "Slide", "Question", "Answer", "Correct", "Question type"])
    ws.append(["u", 1, "Mesmo título", "A", "true", "quiz"])
    ws.append(["u", 2, "Mesmo título", "A", "true", "quiz"]); wb.save(duplicate)
    duplicate_questions, _ = parse_workbook(duplicate, "p")
    assert len({q.question_id for q in duplicate_questions}) == 2


def test_parser_supports_column_oriented_matrix(tmp_path: Path) -> None:
    path = tmp_path / "matrix.xlsx"
    wb = Workbook(); ws = wb.active
    ws.append(["Participant ID", "Q1: Conduta", "Q2: Diagnóstico"])
    ws.append(["alice", "A", "B"]); ws.append(["bob", "B", "A"]); wb.save(path)
    questions, responses = parse_workbook(path, "p")
    assert [q.slide_index for q in questions] == [1, 2]
    assert len(responses) == 4
    keyed = responses_with_answer_keys([
        Question(questions[0].question_id, "p", 1, questions[0].title, QuestionKind.ACADEMIC, ("A", "B"), (0,)),
        Question(questions[1].question_id, "p", 2, questions[1].title, QuestionKind.ACADEMIC, ("A", "B"), (1,)),
    ], responses)
    assert [response.is_correct for response in keyed] == [True, True, False, False]


def test_parser_supports_official_voters_export_with_scales_and_sessions(tmp_path: Path) -> None:
    path = tmp_path / "voters.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Voters"
    sheet.append(["Note: there are more sheets in this document"])
    sheet.append(["Each session is found on its own sheet below."])
    sheet.append([
        "Date (UTC)", "Session", "Voter", "Opening reaction:",
        "Participante: Name", "Participante: Emoji", "Participante: Answer", "Participante: Score",
        "Conduta inicial?: Name", "Conduta inicial?: Emoji", "Conduta inicial?: Answer", "Conduta inicial?: Score",
        "Avalie os seguintes critérios:: Discussão técnica",
        "Avalie os seguintes critérios:: Aplicabilidade",
        "Closing reaction:",
    ])
    for event_date, session in (("2026-05-20", 1), ("2026-05-27", 2)):
        for voter in (1, 2):
            sheet.append([
                event_date, session, voter, "Thumbs up",
                f"name-{session}-{voter}", ":emoji:", "Residente", 0,
                f"name-{session}-{voter}", ":emoji:", "Operar", 900,
                5, 4, "Thumbs up",
            ])
    workbook.save(path)

    questions, responses = parse_workbook(path, "presentation")

    assert [question.title for question in questions] == [
        "Participante",
        "Conduta inicial?",
        "Avalie os seguintes critérios — Discussão técnica",
        "Avalie os seguintes critérios — Aplicabilidade",
    ]
    assert [question.slide_index for question in questions] == [1, 2, 3, 3]
    assert [question.kind for question in questions] == [
        QuestionKind.PROFILE,
        QuestionKind.OTHER,
        QuestionKind.EVALUATION,
        QuestionKind.EVALUATION,
    ]
    assert len(responses) == 16
    assert len({response.session_id for response in responses}) == 2
    assert len({response.participant_id for response in responses}) == 4
    assert valid_response_count(responses, questions) == 12
    assert {response.submitted_at.date() for response in responses} == {
        date(2026, 5, 20), date(2026, 5, 27)
    }
    assert all(response.submitted_at.tzinfo == timezone.utc for response in responses)
    assert not any(response.value in {"Thumbs up", ":emoji:", 0, 900} for response in responses)
    by_session = {
        session_id: [response for response in responses if response.session_id == session_id]
        for session_id in {response.session_id for response in responses}
    }
    assert {
        session_date_from_responses(rows, date(2026, 5, 27))
        for rows in by_session.values()
    } == {date(2026, 5, 20), date(2026, 5, 27)}


def test_parser_preserves_distinct_sessions_from_same_presentation(tmp_path: Path) -> None:
    path = tmp_path / "multiple-sessions.xlsx"
    workbook = Workbook(); sheet = workbook.active
    sheet.append(["Participant ID", "Session ID", "Question", "Answer", "Correct", "Question type"])
    for session in ("morning", "afternoon"):
        for participant in ("alice", "bob"):
            sheet.append([participant, session, "Conduta inicial", "A", "true", "quiz"])
    workbook.save(path)

    _, responses = parse_workbook(path, "presentation")
    assert len({response.session_id for response in responses}) == 2
    assert len({response.response_id for response in responses}) == 4


def test_redaction_is_deterministic_for_common_names() -> None:
    raw = "Nome: João da Silva; Maria Souza e Ricardo Almeida enviaram maria@example.com"
    redacted = redact_pii(raw)
    assert redacted == redact_pii(raw)
    assert "João" not in redacted and "Maria Souza" not in redacted and "Ricardo Almeida" not in redacted and "maria@" not in redacted


def test_redaction_handles_lowercase_context_without_erasing_clinical_terms() -> None:
    redacted = redact_pii("o paciente ricardo almeida apresenta Trauma Abdominal Fechado")
    assert "ricardo almeida" not in redacted
    assert "paciente [NAME]" in redacted
    assert "Trauma Abdominal Fechado" in redacted
    assert "Ricardo Almeida" not in redact_pii("Ricardo Almeida apresenta dor abdominal")
    assert "Maria Clara Souza" not in redact_pii("Maria Clara Souza refere dor")
    assert "Renato Silva" not in redact_pii("Renato Silva apresenta dor")
    assert "Márcio Sousa" not in redact_pii("Márcio Sousa relata dor")
    assert "Gerson Pereira" not in redact_pii("Gerson Pereira apresenta dor")
    assert "renato silva" not in redact_pii("renato silva sofreu queda")
    assert "gerson pereira" not in redact_pii("gerson pereira foi atropelado")
    assert "márcio sousa" not in redact_pii("márcio sousa, 45 anos, apresenta dor")


def test_api_minimization_is_fail_closed_and_preserves_controlled_clinical_terms() -> None:
    minimized = minimize_for_ai(redact_pii(
        "renato silva caiu da moto; Via Aérea Difícil; Controle de Danos; Trauma Abdominal Fechado"
    ))
    assert "renato" not in minimized and "silva" not in minimized
    assert "Via Aérea Difícil" in minimized
    assert "Controle de Danos" in minimized
    assert "Trauma Abdominal Fechado" in minimized


def test_ai_budget_initial_spend_reserve_and_full_taxonomy() -> None:
    class Responses:
        def parse(self, **kwargs):
            assert kwargs["max_output_tokens"] == 500
            return type("R", (), {"output_parsed": Classification(
                primary_topic="Abdome e pelve", subtopic="Baço", cognitive_task="conduta",
                bloom="aplicar", predicted_difficulty="medium", confidence=.9,
                rationale="Classificação clínica", status="classified",
            ), "usage": type("U", (), {"input_tokens": 100, "output_tokens": 100})()})()
    client = type("C", (), {"responses": Responses()})()
    classifier = AIClassifier(client, budget_usd=5, initial_spend_usd=4.9995, input_usd_per_million=1, output_usd_per_million=6, max_output_tokens=500)
    assert classifier.classify("Q", []).status == "pending_budget"


def test_ai_cost_uses_cached_input_rate() -> None:
    class Responses:
        def parse(self, **kwargs):
            usage = type("Usage", (), {
                "input_tokens": 1_000_000,
                "input_tokens_details": type("InputDetails", (), {"cached_tokens": 500_000})(),
                "output_tokens": 100_000,
            })()
            parsed = Classification(
                primary_topic="Abdome e pelve", subtopic="Baço", cognitive_task="conduta",
                bloom="aplicar", predicted_difficulty="medium", confidence=.9,
                rationale="Classificação clínica", status="classified",
            )
            return type("Response", (), {"output_parsed": parsed, "usage": usage})()

    classifier = AIClassifier(
        type("Client", (), {"responses": Responses()})(), budget_usd=10,
        input_usd_per_million=1, cached_input_usd_per_million=.1,
        output_usd_per_million=6, max_output_tokens=500,
    )
    classifier.classify("Q", [])
    assert classifier.usage_summary["cached_input_tokens"] == 500_000
    assert classifier.usage_summary["estimated_cost_usd"] == pytest.approx(1.15)


def test_incremental_reports_use_full_repository_corpus(tmp_path: Path, synthetic_reference_xlsx: Path, synthetic_slide_deck: dict) -> None:
    repository = LocalRepository(tmp_path / "repo")
    old_p = Presentation("old", "Desafio Trauma - 20/05/2026", date(2026, 5, 20), "/old")
    old_s = Session("old-s", "old", date(2026, 5, 20), 5, 1, True)
    old_q = Question("old-q", "old", 1, "Old", QuestionKind.ACADEMIC, ("A", "B"), (0,))
    old_r = [Response(f"old-r{i}", "old-s", "old-q", f"old-u{i}", "A", True) for i in range(5)]
    repository._corpus = ([old_p], [old_s], [old_q], old_r)

    class Client:
        def discover(self): return [PresentationRef("new", "Desafio Trauma - 27/05/2026", "/new")]
        def fetch(self, ref, destination): destination.mkdir(parents=True, exist_ok=True); return synthetic_reference_xlsx, synthetic_slide_deck

    result = Pipeline(client=Client(), repository=repository, workdir=tmp_path / "work").sync("incremental")
    assert result["presentations"] == 1 and result["corpus_presentations"] == 2
    private_report = next((tmp_path / "work/reports").glob("*/private.xlsx"))
    assert load_workbook(private_report, read_only=True)["Sessões"].max_row == 3
    assert json.loads((tmp_path / "repo/last_good_snapshot.json").read_text())["snapshot"]["overview"]["presentations"] == 2


def test_partial_sessions_and_empty_copies_do_not_enter_publication(tmp_path: Path) -> None:
    presentation = Presentation("p", "Desafio Trauma - 22/07/2026", date(2026, 7, 22), "/p")
    sessions = [Session("closed", "p", date(2026, 7, 15), 5, 1, True), Session("live", "p", date(2026, 7, 22), 5, 1, False)]
    question = Question("q", "p", 1, "Q", QuestionKind.ACADEMIC)
    responses = [Response(f"closed-{i}", "closed", "q", f"c{i}", "A", True) for i in range(5)] + [Response(f"live-{i}", "live", "q", f"l{i}", "A", True) for i in range(5)]
    _, published_sessions, _, published_responses = published_corpus([presentation], sessions, [question], responses)
    assert [session.session_id for session in published_sessions] == ["closed"]
    assert {response.session_id for response in published_responses} == {"closed"}

    empty = tmp_path / "empty-copy.xlsx"
    workbook = Workbook(); workbook.active.append(["Participant ID", "Question", "Answer"]); workbook.save(empty)
    repository = LocalRepository(tmp_path / "repo")
    repository._corpus = ([presentation], [sessions[0]], [question], responses[:5])
    class Client:
        def discover(self): return [PresentationRef("empty", "Desafio Trauma - 22/07/2026", "/empty")]
        def fetch(self, ref, destination): return empty, {"slides": []}
    result = Pipeline(client=Client(), repository=repository, workdir=tmp_path / "work").sync("incremental")
    assert result["presentations"] == 0
    assert json.loads((tmp_path / "repo/last_good_snapshot.json").read_text())["snapshot"]["overview"]["presentations"] == 1


def test_manual_without_id_processes_pending_import(tmp_path: Path, synthetic_reference_xlsx: Path) -> None:
    class ManualRepository(LocalRepository):
        def pending_manual_imports(self, destination):
            return [ManualImport("import-1", synthetic_reference_xlsx, "manual-p", date(2026, 5, 27), "Desafio Trauma - 27/05/2026")]

    class NoDiscovery:
        def discover(self): raise AssertionError("manual imports must not call discovery")

    result = Pipeline(client=NoDiscovery(), repository=ManualRepository(tmp_path / "repo"), workdir=tmp_path / "work").sync("manual")
    assert result["responses"] == 208


def test_empty_manual_import_is_rejected_and_never_marked_imported(tmp_path: Path) -> None:
    empty = tmp_path / "empty.xlsx"
    workbook = Workbook(); workbook.active.append(["Participant ID", "Question", "Answer"]); workbook.save(empty)

    class ManualRepository(LocalRepository):
        rejected: list[tuple[str, str]] = []

        def pending_manual_imports(self, destination):
            return [ManualImport("import-empty", empty, "manual-empty", date(2026, 5, 27), "Desafio Trauma - 27/05/2026")]

        def reject_manual_import(self, import_id: str, reason: str) -> None:
            self.rejected.append((import_id, reason))

        def publish_snapshot(self, *args, **kwargs):
            raise AssertionError("an empty manual import must not publish a snapshot")

    repository = ManualRepository(tmp_path / "repo")
    with pytest.raises(RuntimeError, match="No valid manual imports"):
        Pipeline(client=object(), repository=repository, workdir=tmp_path / "work").sync("manual")
    assert repository.rejected == [("import-empty", "Arquivo reconhecido, mas sem respostas válidas para importação.")]


def test_unknown_manual_schema_is_rejected_without_poisoning_valid_imports(tmp_path: Path, synthetic_reference_xlsx: Path) -> None:
    unknown = tmp_path / "unknown.xlsx"
    workbook = Workbook(); workbook.active.append(["Coluna arbitrária"]); workbook.active.append(["valor"]); workbook.save(unknown)

    class ManualRepository(LocalRepository):
        rejected: list[tuple[str, str]] = []

        def pending_manual_imports(self, destination):
            return [
                ManualImport("bad", unknown, "manual-bad", date(2026, 5, 20), "Desafio Trauma - 20/05/2026"),
                ManualImport("good", synthetic_reference_xlsx, "manual-good", date(2026, 5, 27), "Desafio Trauma - 27/05/2026"),
            ]

        def reject_manual_import(self, import_id: str, reason: str) -> None:
            self.rejected.append((import_id, reason))

    repository = ManualRepository(tmp_path / "repo")
    result = Pipeline(client=object(), repository=repository, workdir=tmp_path / "work").sync("manual")
    assert result["presentations"] == 1
    assert repository.rejected[0][0] == "bad"
    assert "não reconhecida" in repository.rejected[0][1]


def test_corrupt_manual_xlsx_is_rejected_without_poisoning_valid_imports(tmp_path: Path, synthetic_reference_xlsx: Path) -> None:
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not a zip archive")

    class ManualRepository(LocalRepository):
        rejected: list[tuple[str, str]] = []

        def pending_manual_imports(self, destination):
            return [
                ManualImport("corrupt", corrupt, "manual-corrupt", date(2026, 5, 20), "Desafio Trauma - 20/05/2026"),
                ManualImport("good", synthetic_reference_xlsx, "manual-good", date(2026, 5, 27), "Desafio Trauma - 27/05/2026"),
            ]

        def reject_manual_import(self, import_id: str, reason: str) -> None:
            self.rejected.append((import_id, reason))

    repository = ManualRepository(tmp_path / "repo")
    result = Pipeline(client=object(), repository=repository, workdir=tmp_path / "work").sync("manual")
    assert result["presentations"] == 1
    assert repository.rejected[0][0] == "corrupt"


def test_incomplete_ooxml_zip_is_rejected_without_poisoning_queue(tmp_path: Path, synthetic_reference_xlsx: Path) -> None:
    incomplete = tmp_path / "incomplete.xlsx"
    with ZipFile(incomplete, "w") as archive:
        archive.writestr("xl/workbook.xml", "<workbook />")

    class ManualRepository(LocalRepository):
        rejected: list[tuple[str, str]] = []

        def pending_manual_imports(self, destination):
            return [
                ManualImport("incomplete", incomplete, "manual-incomplete", date(2026, 5, 20), "Desafio Trauma - 20/05/2026"),
                ManualImport("good", synthetic_reference_xlsx, "manual-good", date(2026, 5, 27), "Desafio Trauma - 27/05/2026"),
            ]

        def reject_manual_import(self, import_id: str, reason: str) -> None:
            self.rejected.append((import_id, reason))

    repository = ManualRepository(tmp_path / "repo")
    result = Pipeline(client=object(), repository=repository, workdir=tmp_path / "work").sync("manual")
    assert result["presentations"] == 1
    assert repository.rejected[0][0] == "incomplete"


def test_lazy_worksheet_parse_error_is_normalized_and_workbook_closed(tmp_path: Path) -> None:
    malformed = tmp_path / "malformed.xlsx"
    workbook = Workbook(); sheet = workbook.active
    sheet.append(["Participant ID", "Question", "Answer"])
    sheet.append([123, "Q", "A"])
    workbook.save(malformed)
    with ZipFile(malformed, "r") as archive:
        members = {item.filename: archive.read(item.filename) for item in archive.infolist()}
    worksheet_name = "xl/worksheets/sheet1.xml"
    members[worksheet_name] = members[worksheet_name].replace(b"<v>123</v>", b"<v>not-a-number</v>")
    with ZipFile(malformed, "w") as archive:
        for name, content in members.items():
            archive.writestr(name, content)
    with pytest.raises(InvalidWorkbookError):
        parse_workbook(malformed, "p")


def test_invalid_shared_string_reference_is_normalized(tmp_path: Path) -> None:
    malformed = tmp_path / "bad-shared-string.xlsx"
    workbook = Workbook(); sheet = workbook.active
    sheet.append(["Participant ID", "Question", "Answer"])
    sheet.append([123, "Q", "A"])
    workbook.save(malformed)
    with ZipFile(malformed, "r") as archive:
        members = {item.filename: archive.read(item.filename) for item in archive.infolist()}
    worksheet_name = "xl/worksheets/sheet1.xml"
    members[worksheet_name] = members[worksheet_name].replace(
        b'<c r="A2" t="n"><v>123</v></c>',
        b'<c r="A2" t="s"><v>999</v></c>',
    )
    with ZipFile(malformed, "w") as archive:
        for name, content in members.items():
            archive.writestr(name, content)
    with pytest.raises(InvalidWorkbookError):
        parse_workbook(malformed, "p")


def test_classification_is_reused_unless_forced(tmp_path: Path) -> None:
    existing = Question("q", "p", 1, "Q", QuestionKind.ACADEMIC, topic="Abdome e pelve", subtopic="Baço", cognitive_task="conduta", bloom="aplicar", predicted_difficulty="medium", ai_confidence=.9)
    incoming = Question("q", "p", 1, "Q", QuestionKind.ACADEMIC)
    class Classifier:
        calls = 0
        def classify(self, *_): self.calls += 1; return Classification(primary_topic="Tórax", subtopic="Pulmão", cognitive_task="diagnóstico", bloom="aplicar", predicted_difficulty="hard", confidence=.95, rationale="Teste")
    classifier = Classifier(); pipeline = Pipeline(client=object(), repository=LocalRepository(tmp_path), classifier=classifier, workdir=tmp_path)
    assert pipeline._classify(incoming, False, existing).topic == "Abdome e pelve"
    assert classifier.calls == 0
    assert pipeline._classify(incoming, True, existing).topic == "Tórax"


def test_classification_survives_without_api_and_pending_budget_retries(tmp_path: Path) -> None:
    existing = Question("q", "p", 1, "Q", QuestionKind.ACADEMIC, topic="Tórax", ai_confidence=.9, ai_status="reviewed")
    incoming = Question("q", "p", 1, "Q", QuestionKind.ACADEMIC)
    no_api = Pipeline(client=object(), repository=LocalRepository(tmp_path / "none"), classifier=None, workdir=tmp_path / "none")
    assert no_api._classify(incoming, False, existing).topic == "Tórax"

    pending = Question("q", "p", 1, "Q", QuestionKind.ACADEMIC, topic="Outros", ai_confidence=0, ai_status="pending_budget")
    class Classifier:
        calls = 0
        def classify(self, *_):
            self.calls += 1
            return Classification(primary_topic="Tórax", subtopic="Pleura", cognitive_task="conduta", bloom="aplicar", predicted_difficulty="medium", confidence=.9, rationale="Nova tentativa")
    classifier = Classifier()
    with_api = Pipeline(client=object(), repository=LocalRepository(tmp_path / "api"), classifier=classifier, workdir=tmp_path / "api")
    assert with_api._classify(incoming, False, pending).topic == "Tórax"
    assert classifier.calls == 1


def test_public_report_suppresses_accuracy_with_fewer_than_five_people(tmp_path: Path) -> None:
    presentation = Presentation("p", "Desafio Trauma - 01/01/2026", date(2026, 1, 1), "/p")
    session = Session("s", "p", date(2026, 1, 1), 5, 5, True)
    questions = [Question(f"q{i}", "p", i, f"Q{i}", QuestionKind.ACADEMIC) for i in range(5)]
    responses = [Response(f"r{i}", "s", f"q{i}", "same-person", "A", True) for i in range(5)]
    path = write_report(tmp_path / "public.xlsx", [presentation], [session], questions, responses, public=True)
    values = dict(load_workbook(path, data_only=True)["Resumo"].iter_rows(min_row=2, values_only=True))
    assert values["Acurácia acadêmica"] is None


def test_supabase_fetch_all_paginates_past_default_limit() -> None:
    source = [{"id": index} for index in range(1001)]
    class Query:
        def __init__(self): self.start = 0; self.end = 999
        def select(self, *_): return self
        def order(self, *_): return self
        def range(self, start, end): self.start, self.end = start, end; return self
        def execute(self): return type("Response", (), {"data": source[self.start:self.end + 1]})()
    class Client:
        def table(self, *_): return Query()
    repository = SupabaseRepository.__new__(SupabaseRepository); repository.client = Client()
    assert len(repository._fetch_all("table", "id", ("id",))) == 1001


def test_publication_uses_transactional_rpc_as_single_call() -> None:
    captured = {}
    class RPC:
        def execute(self): return None
    class Client:
        def rpc(self, name, args): captured.update({"name": name, "args": args}); return RPC()
    repository = SupabaseRepository.__new__(SupabaseRepository); repository.client = Client(); repository.run_id = "00000000-0000-0000-0000-000000000001"
    repository.publish_snapshot("ignored", "public.xlsx", "private.xlsx", {"ok": True}, {"selected": 1}, {"input_tokens": 2, "cached_input_tokens": 1, "output_tokens": 3, "estimated_cost_usd": .01})
    assert captured["name"] == "publish_dashboard_snapshot"
    assert set(captured["args"]) == {"p_pipeline_run_id", "p_schema_version", "p_snapshot", "p_privacy_k", "p_checksum_sha256", "p_result", "p_manual_imports", "p_input_tokens", "p_cached_input_tokens", "p_output_tokens", "p_estimated_cost_usd"}


def test_question_classification_uses_dedicated_schema_columns() -> None:
    question = Question("q", "p", 1, "Q", QuestionKind.ACADEMIC, ("A",), (0,), topic="Abdome e pelve", analysis_role="academic", subtopic="Baço", cognitive_task="conduta", bloom="aplicar", predicted_difficulty="medium", ai_confidence=.9, ai_rationale="Razão clínica", ai_status="classified", taxonomy_version="v1")
    payload = question_payload(question, "db-p", 1)
    assert payload["slide_index"] == 1
    assert payload["primary_topic"] == "Abdome e pelve"
    assert payload["bloom_level"] == "aplicar"
    assert payload["ai_status"] == "classified"
    assert all("_meta" not in option for option in payload["options"])
