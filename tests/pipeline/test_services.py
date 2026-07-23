from __future__ import annotations

import json
from datetime import date, datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from pipeline.ai import AIClassifier, Classification, redact_pii
from pipeline.mentimeter import (
    PresentationRef,
    extract_latest_slide_deck,
    extract_slide_deck,
    matches_title,
    remember_json_response,
    select_presentations,
)
from pipeline.models import Presentation, Question, QuestionKind, Response, Session
from pipeline.orchestrator import Pipeline, SnapshotManager, questions_from_deck
from pipeline.persistence import LocalRepository
from pipeline.reports import PRIVATE_SHEETS, PUBLIC_SHEETS, write_report


def test_discovery_title_is_strict_and_selection_reprocesses_two_recent() -> None:
    assert matches_title("Desafio Trauma - 27/05/2026")
    assert matches_title("Desafio Trauma -27/05/2026")
    assert not matches_title("Desafio Trauma - 27/05/2026 copy")
    refs = [
        PresentationRef(str(i), f"Desafio Trauma - {i:02d}/05/2026", f"/{i}", complete=True)
        for i in range(1, 5)
    ]
    refs[0] = PresentationRef("1", refs[0].title, "/1", complete=False)
    selected = select_presentations(refs, mode="incremental", known_ids={r.presentation_id for r in refs})
    assert {r.presentation_id for r in selected} == {"1", "3", "4"}
    older = PresentationRef("legacy", "Desafio Trauma - 22/10/2024", "/legacy")
    assert select_presentations([older, *refs], mode="backfill") == refs


def test_mentimeter_credentials_support_local_fallback(monkeypatch) -> None:
    from pipeline.mentimeter import MentimeterClient

    monkeypatch.delenv("MENTIMETER_EMAIL", raising=False)
    monkeypatch.delenv("MENTIMETER_PASSWORD", raising=False)
    monkeypatch.setenv("LOGIN_EMAIL", "local@example.invalid")
    monkeypatch.setenv("LOGIN_PASSWORD", "not-a-real-secret")
    client = MentimeterClient()
    assert client.email.endswith(".invalid")


def test_mentimeter_reuses_ephemeral_authenticated_storage_state(tmp_path: Path, monkeypatch) -> None:
    from pipeline.mentimeter import MentimeterClient

    monkeypatch.setenv("PIPELINE_WORKDIR", str(tmp_path))

    class FakePage:
        pass

    class FakeContext:
        def __init__(self) -> None:
            self.page = FakePage()

        def new_page(self):
            return self.page

        def storage_state(self, *, path: str) -> None:
            Path(path).write_text("{}", encoding="utf-8")

    class FakeBrowser:
        def __init__(self) -> None:
            self.options = []

        def new_context(self, **options):
            self.options.append(options)
            return FakeContext()

    client = MentimeterClient(email="admin@example.invalid", password="secret")
    logins = []
    client._login = lambda page: logins.append(page)
    browser = FakeBrowser()

    client._authenticated_page(browser)
    client._authenticated_page(browser)

    assert len(logins) == 1
    assert browser.options == [
        {"accept_downloads": True},
        {
            "accept_downloads": True,
            "storage_state": str(tmp_path / "mentimeter-storage-state.json"),
        },
    ]


def test_mentimeter_login_uses_stable_form_test_ids() -> None:
    from pipeline.mentimeter import MentimeterClient

    class FakeLocator:
        def __init__(self) -> None:
            self.value = None
            self.clicked = False
            self.click_options = {}
            self.removed = False

        def fill(self, value: str) -> None:
            self.value = value

        def click(self, **options) -> None:
            self.clicked = True
            self.click_options = options

        def count(self) -> int:
            return 1

        def evaluate_all(self, _script: str) -> None:
            self.removed = True

    class FakePage:
        def __init__(self) -> None:
            self.email = FakeLocator()
            self.password = FakeLocator()
            self.submit = FakeLocator()
            self.consent = FakeLocator()
            self.test_ids: list[str] = []

        def goto(self, *_args, **_kwargs) -> None:
            return None

        def get_by_label(self, pattern):
            if pattern.search("Show password"):
                raise AssertionError("password label selector also matches the visibility button")
            return self.email

        def get_by_test_id(self, test_id: str):
            self.test_ids.append(test_id)
            return {"password-input": self.password, "login-btn": self.submit}[test_id]

        def locator(self, selector: str):
            assert selector == "#cookiebanner, #cookiebanner-container, #cookiebanner-backdrop"
            return self.consent

        def get_by_role(self, _role, name):
            if name.search("Sign in with Google"):
                raise AssertionError("login role selector also matches social auth buttons")
            return self.submit

        def wait_for_url(self, *_args, **_kwargs) -> None:
            return None

    page = FakePage()
    MentimeterClient(email="admin@example.invalid", password="secret")._login(page)

    assert page.test_ids == ["password-input", "login-btn"]
    assert page.email.value == "admin@example.invalid"
    assert page.password.value == "secret"
    assert page.consent.removed is True
    assert page.submit.clicked is True
    assert page.submit.click_options == {}


def test_mentimeter_discovery_uses_named_folder_and_waits_for_all_cards() -> None:
    from pipeline.mentimeter import MentimeterClient

    class FakeLocator:
        def __init__(self, items) -> None:
            self.items = items

        def count(self) -> int:
            return len(self.items)

        def nth(self, index: int):
            return self.items[index]

    class FakeLink:
        def __init__(self, text: str, href: str) -> None:
            self.text = text
            self.href = href

        def inner_text(self) -> str:
            return self.text

        def get_attribute(self, name: str):
            return self.href if name == "href" else None

    class FakePage:
        def __init__(self) -> None:
            self.urls: list[str] = []
            self.states = iter([
                {"links": 3, "loading": True, "scrollHeight": 1000},
                {"links": 6, "loading": False, "scrollHeight": 2000},
                {"links": 6, "loading": False, "scrollHeight": 2000},
                {"links": 6, "loading": False, "scrollHeight": 2000},
                {"links": 6, "loading": False, "scrollHeight": 2000},
            ])

        def goto(self, url: str, **_kwargs) -> None:
            self.urls.append(url)

        def wait_for_selector(self, _selector: str, **_kwargs) -> None:
            return None

        def locator(self, selector: str):
            if "folder" in selector:
                return FakeLocator([
                    FakeLink("Other", "/app/folder/1"),
                    FakeLink("Desafio Trauma", "/app/folder/2601315"),
                ])
            assert 'href*="/edit"' in selector
            return FakeLocator([
                FakeLink("Desafio Trauma - 06/11/2024", "/app/presentation/p1/edit?source=dashboard"),
                FakeLink("Edited November 6, 2024", "/app/presentation/p1/edit?source=dashboard"),
                FakeLink("Desafio Trauma - 06/11/2024", "/app/presentation/p1/edit?source=dashboard"),
                FakeLink("Desafio Trauma - 22/07/2026 copy (1)", "/app/presentation/copy/edit"),
                FakeLink("Desafio Trauma - 22/07/2026", "/app/presentation/p2/edit?source=dashboard"),
                FakeLink("Desafio Trauma - 22/07/2026", "/app/presentation/p2/edit?source=dashboard"),
            ])

        def evaluate(self, _script: str):
            return next(self.states)

        def wait_for_timeout(self, _milliseconds: int) -> None:
            return None

    page = FakePage()
    refs = MentimeterClient(email="admin@example.invalid", password="secret")._discover_with_page(page)

    assert page.urls == [
        "https://www.mentimeter.com/app/dashboard",
        "https://www.mentimeter.com/app/folder/2601315",
    ]
    assert [(ref.presentation_id, ref.title) for ref in refs] == [
        ("p1", "Desafio Trauma - 06/11/2024"),
        ("p2", "Desafio Trauma - 22/07/2026"),
    ]
    assert [ref.href for ref in refs] == [
        "/app/presentation/p1/results?source=dashboard",
        "/app/presentation/p2/results?source=dashboard",
    ]


def test_mentimeter_download_uses_results_page_and_xlsx_menuitem(tmp_path: Path) -> None:
    from pipeline.mentimeter import MentimeterClient

    class FakeLocator:
        def __init__(self, fail_first_wait: bool = False) -> None:
            self.waited = False
            self.wait_calls = 0
            self.fail_first_wait = fail_first_wait
            self.click_options = None
            self.removed = False

        def wait_for(self, **_kwargs) -> None:
            self.waited = True
            self.wait_calls += 1
            if self.fail_first_wait and self.wait_calls == 1:
                raise TimeoutError("simulated delayed results controls")

        def click(self, **options) -> None:
            self.click_options = options

        def count(self) -> int:
            return 1

        def evaluate_all(self, _script: str) -> None:
            self.removed = True

    class FakeDownload:
        def save_as(self, destination: Path) -> None:
            destination.write_bytes(b"xlsx")

    class FakeDownloadContext:
        value = FakeDownload()

        def __init__(self, fail: bool = False) -> None:
            self.fail = fail

        def __enter__(self):
            return self

        def __exit__(self, *_args) -> None:
            if self.fail:
                raise TimeoutError("simulated delayed XLSX generation")
            return None

    class FakePage:
        def __init__(self) -> None:
            self.goto_call = None
            self.download_button = FakeLocator(fail_first_wait=True)
            self.xlsx_menuitem = FakeLocator()
            self.consent = FakeLocator()
            self.roles = []
            self.reload_calls = []
            self.expect_calls = 0

        def goto(self, *args, **kwargs) -> None:
            self.goto_call = (args, kwargs)

        def get_by_role(self, role: str, **options):
            self.roles.append((role, options))
            assert role == "button"
            return self.download_button

        def reload(self, **options) -> None:
            self.reload_calls.append(options)

        def locator(self, selector: str):
            if selector == "#cookiebanner, #cookiebanner-container, #cookiebanner-backdrop":
                return self.consent
            assert selector == "#excel-download-button"
            return self.xlsx_menuitem

        def expect_event(self, event: str, **_kwargs):
            assert event == "download"
            self.expect_calls += 1
            return FakeDownloadContext(fail=self.expect_calls == 1)

    page = FakePage()
    ref = PresentationRef("p1", "Desafio Trauma - 27/05/2026", "/app/presentation/p1/results?source=dashboard")
    path = MentimeterClient(email="admin@example.invalid", password="secret")._download_with_page(
        page, ref, tmp_path
    )

    assert page.goto_call == (
        ("https://www.mentimeter.com/app/presentation/p1/results?source=dashboard",),
        {"wait_until": "domcontentloaded", "timeout": 45_000},
    )
    assert page.roles[0] == ("button", {"name": "Download", "exact": True})
    assert page.reload_calls == [
        {"wait_until": "domcontentloaded", "timeout": 45_000},
        {"wait_until": "domcontentloaded", "timeout": 45_000},
    ]
    assert page.consent.removed is True
    assert page.download_button.waited is True
    assert page.download_button.click_options == {}
    assert page.xlsx_menuitem.waited is True
    assert page.xlsx_menuitem.click_options == {}
    assert path.read_bytes() == b"xlsx"


def test_capture_accepts_only_json_with_slide_deck() -> None:
    assert extract_slide_deck({"anything": 1}) is None
    deck = {"slide_deck": {"name": "Desafio Trauma - 27/05/2026", "slides": []}}
    assert extract_slide_deck(deck) == deck["slide_deck"]


def test_response_callback_defers_body_read_until_page_operation_finishes() -> None:
    class FakeResponse:
        headers = {"content-type": "application/json; charset=utf-8"}

        def __init__(self) -> None:
            self.json_calls = 0

        def json(self):
            self.json_calls += 1
            return {"data": {"slide_deck": {"name": "Desafio Trauma - 27/05/2026", "slides": []}}}

    response = FakeResponse()
    candidates = []
    remember_json_response(candidates, response)
    assert response.json_calls == 0

    deck = extract_latest_slide_deck(candidates)
    assert response.json_calls == 1
    assert deck["name"] == "Desafio Trauma - 27/05/2026"


def test_slide_deck_builds_academic_question_from_correct_markers() -> None:
    deck = {"slides": [{"static_content": {"type": "quiz-choice"}, "interactive_contents": [{
        "title": "Conduta?", "correct_answer_mode": "enabled",
        "choices": [{"title": "A", "marked_correct": False}, {"title": "B", "marked_correct": True}],
    }]}]}
    question = questions_from_deck(deck, "p1")[0]
    assert question.kind == QuestionKind.ACADEMIC
    assert question.correct_indices == (1,)


def test_dry_run_only_discovers_and_selects(tmp_path: Path) -> None:
    class FakeClient:
        def discover(self):
            return [PresentationRef("p1", "Desafio Trauma - 27/05/2026", "/p1")]

        def fetch(self, *_):
            raise AssertionError("dry run must not download")

    result = Pipeline(client=FakeClient(), repository=LocalRepository(tmp_path), workdir=tmp_path).sync(
        "incremental", dry_run=True
    )
    assert result["selected"] == 1


def test_local_pipeline_end_to_end_creates_last_good_only_after_publish(
    tmp_path: Path, synthetic_reference_xlsx: Path, synthetic_slide_deck: dict
) -> None:
    class FakeClient:
        def discover(self):
            return [PresentationRef("p1", "Desafio Trauma - 27/05/2026", "/p1")]

        def fetch(self, ref, destination):
            destination.mkdir(parents=True, exist_ok=True)
            return synthetic_reference_xlsx, synthetic_slide_deck

    repository = LocalRepository(tmp_path / "repository")
    result = Pipeline(client=FakeClient(), repository=repository, workdir=tmp_path / "work").sync("backfill")
    assert result["responses"] == 208
    assert len(list((tmp_path / "work/reports").glob("*/public.xlsx"))) == 1
    published = json.loads((tmp_path / "repository/last_good_snapshot.json").read_text())["snapshot"]
    assert set(published) == {"metadata", "filters", "overview", "participation", "learning", "experience", "topics", "metric_dictionary", "public_files"}
    assert published["overview"]["response_rate"] == 84.9  # 208 / (35 * 7)
    assert published["learning"]["accuracy_rate"] == 85.5  # 142 / 166
    assert published["experience"]["evaluations"] == 9      # 9 / 35 evaluated
    assert published["experience"]["evaluation_rate"] == 25.7


def test_ai_redacts_pii_uses_structured_parse_and_flags_low_confidence() -> None:
    captured = {}

    class FakeResponses:
        def parse(self, **kwargs):
            captured.update(kwargs)
            return type("Result", (), {"output_parsed": Classification(
                primary_topic="Abdome e pelve", subtopic="Baço", cognitive_task="conduta",
                bloom="aplicar", predicted_difficulty="medium", confidence=0.79,
                rationale="Classificação clínica",
            ), "usage": type("Usage", (), {"input_tokens": 100, "output_tokens": 20})()})()

    client = type("Client", (), {"responses": FakeResponses()})()
    classifier = AIClassifier(client=client, budget_usd=5)
    result = classifier.classify("João Silva joao@example.com CPF 123.456.789-00", ["A", "B"])
    assert result.needs_review is True and result.primary_topic == "Abdome e pelve"
    assert captured["model"] == "gpt-5.6-luna"
    assert captured["store"] is False
    assert captured["reasoning"] == {"effort": "low"}
    prompt = captured["input"]
    assert "joao@example.com" not in prompt and "123.456.789-00" not in prompt
    assert "[EMAIL]" in redact_pii("a@b.com")


def test_ai_budget_warns_and_stops() -> None:
    classifier = AIClassifier(client=object(), budget_usd=5)
    classifier.spent_usd = 3.5
    assert classifier.budget_warning
    classifier.spent_usd = 5
    assert classifier.classify("Q", []) .status == "pending_budget"


def test_ai_timeout_marks_question_for_review_without_blocking_pipeline() -> None:
    class TimeoutResponses:
        def parse(self, **_kwargs):
            raise TimeoutError("simulated API timeout")

    client = type("Client", (), {"responses": TimeoutResponses()})()
    classifier = AIClassifier(client=client, budget_usd=5)

    result = classifier.classify("Questão clínica", ["A", "B"])

    assert result.status == "failed"
    assert result.needs_review is True
    assert result.primary_topic == "Outros"
    assert classifier.usage_summary["estimated_cost_usd"] == 0


def _sample_data():
    presentation = Presentation("p1", "Desafio Trauma - 27/05/2026", date(2026, 5, 27), "/p1")
    session = Session("s1", "p1", date(2026, 5, 27), 5, 1, complete=True)
    question = Question("q1", "p1", 1, "Pergunta", QuestionKind.ACADEMIC, ("A", "B"), (0,))
    responses = [Response(f"r{i}", "s1", "q1", f"u{i}", "A", True) for i in range(5)]
    return [presentation], [session], [question], responses


def test_reports_have_exact_sheets_and_public_suppression(tmp_path: Path) -> None:
    data = _sample_data()
    public = tmp_path / "public.xlsx"
    private = tmp_path / "private.xlsx"
    write_report(public, *data, public=True)
    write_report(private, *data, public=False)
    public_wb = load_workbook(public, read_only=True)
    assert public_wb.sheetnames == list(PUBLIC_SHEETS)
    assert load_workbook(private, read_only=True).sheetnames == list(PRIVATE_SHEETS)
    assert public_wb["Questões agregadas"]["F2"].value == 1
    assert "Correlação ponto-bisserial" in [cell.value for cell in public_wb["Questões agregadas"][1]]
    assert public_wb["Encontros"]["F2"].value == 1


def test_snapshot_only_replaces_last_good_after_success(tmp_path: Path) -> None:
    manager = SnapshotManager(tmp_path)
    manager.publish({"version": 1})
    assert json.loads((tmp_path / "last_good.json").read_text())["version"] == 1
    with pytest.raises(RuntimeError):
        manager.publish({"version": 2}, validate=lambda _: (_ for _ in ()).throw(RuntimeError("bad")))
    assert json.loads((tmp_path / "last_good.json").read_text())["version"] == 1
