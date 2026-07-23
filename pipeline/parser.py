from __future__ import annotations

import re
import unicodedata
from zipfile import BadZipFile
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import ParseError

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .ids import content_hash, stable_id
from .models import Question, QuestionKind, Response


class UnknownSchemaError(ValueError):
    """Raised when no known Mentimeter export signature can be identified."""


class EmptyPresentationError(UnknownSchemaError):
    """Raised for a recognized export with no participant responses."""


class InvalidWorkbookError(UnknownSchemaError):
    """Raised when an XLSX container cannot be read safely."""


ALIASES = {
    "participant": {"participant id", "participant", "id do participante", "participante id", "respondent id"},
    "question": {"question", "question title", "pergunta", "questao", "titulo da pergunta"},
    "answer": {"answer", "response", "resposta", "respostas"},
    "correct": {"correct", "is correct", "correta", "correto", "acertou"},
    "type": {"question type", "slide type", "tipo", "tipo de pergunta"},
    "submitted": {"submitted at", "timestamp", "data", "respondido em"},
    "slide": {"slide", "slide number", "numero do slide", "n do slide", "question number"},
    "session": {"session", "session id", "sessao", "id da sessao", "identificador da sessao"},
}


def _key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().casefold()


def _mapping(row: tuple[Any, ...]) -> dict[str, int]:
    normalized = [_key(value) for value in row]
    result: dict[str, int] = {}
    for field, aliases in ALIASES.items():
        aliases_normalized = {_key(alias) for alias in aliases}
        for index, value in enumerate(normalized):
            if value in aliases_normalized:
                result[field] = index
                break
    return result


def _kind(title: str, raw_type: str) -> QuestionKind:
    sample = _key(f"{raw_type} {title}")
    if any(token in sample for token in ("participante", "participant", "perfil", "profile")):
        return QuestionKind.PROFILE
    if "nps" in sample or "recomend" in sample:
        return QuestionKind.NPS
    if any(token in sample for token in ("avaliacao", "evaluation", "rating", "scale", "escala")):
        return QuestionKind.EVALUATION
    if any(token in sample for token in ("quiz", "choice", "escolha")):
        return QuestionKind.ACADEMIC
    return QuestionKind.OTHER


def _boolean(value: Any) -> bool | None:
    token = _key(value)
    if token in {"true", "yes", "sim", "correct", "correta", "correto", "1"}:
        return True
    if token in {"false", "no", "nao", "incorrect", "incorreta", "incorreto", "0"}:
        return False
    return None


def parse_workbook(path: str | Path, presentation_id: str) -> tuple[list[Question], list[Response]]:
    """Parse known row-oriented pt/en export shapes, failing closed otherwise."""
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        return _parse_loaded_workbook(workbook, path, presentation_id)
    except (EmptyPresentationError, UnknownSchemaError):
        raise
    except (BadZipFile, InvalidFileException, ParseError, EOFError, IndexError, KeyError, OSError, ValueError) as exc:
        raise InvalidWorkbookError(f"Unreadable XLSX workbook: {Path(path).name}") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _parse_loaded_workbook(workbook, path: str | Path, presentation_id: str) -> tuple[list[Question], list[Response]]:
    selected = None
    matrix = None
    voters = None
    for sheet in workbook.worksheets:
        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            normalized = [_key(value) for value in row]
            if (
                "date (utc)" in normalized
                and "session" in normalized
                and "voter" in normalized
                and any(re.search(r":\s*answer$", value, re.I) for value in normalized)
            ):
                voters = voters or (sheet, row_number, row)
            mapping = _mapping(row)
            if {"participant", "question", "answer"}.issubset(mapping):
                selected = (sheet, row_number, mapping)
                break
            if "participant" in mapping and len([value for value in row if value not in (None, "")]) >= 2:
                matrix = matrix or (sheet, row_number, mapping["participant"], row)
        if selected:
            break
    if selected is None and voters is not None:
        return _parse_voters_export(*voters, presentation_id, path)
    if selected is None and matrix is not None:
        return _parse_matrix(*matrix, presentation_id)
    if selected is None:
        raise UnknownSchemaError(f"Unknown XLSX schema: {Path(path).name}")

    sheet, header_row, mapping = selected
    questions_by_key: dict[tuple[int, str], Question] = {}
    responses_by_id: dict[str, Response] = {}
    fallback_slide_by_title: dict[str, int] = {}
    for excel_row, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        def get(field: str) -> Any:
            index = mapping.get(field)
            return row[index] if index is not None and index < len(row) else None

        title = str(get("question") or "").strip()
        answer = get("answer")
        if not title or answer is None or str(answer).strip() == "":
            continue
        raw_type = str(get("type") or "")
        participant = str(get("participant") or "").strip()
        if not participant:
            continue
        slide_raw = get("slide")
        try:
            if slide_raw not in (None, ""):
                slide_index = int(slide_raw)
                slide_identity: int | str = slide_index
            else:
                slide_index = fallback_slide_by_title.setdefault(title, len(fallback_slide_by_title) + 1)
                slide_identity = f"title:{title}"
        except (TypeError, ValueError):
            slide_index = len(questions_by_key) + 1
            slide_identity = f"title:{title}"
        question_id = stable_id("question", presentation_id, slide_identity, title)
        key = (slide_index, title)
        if key not in questions_by_key:
            questions_by_key[key] = Question(
                question_id=question_id, presentation_id=presentation_id, slide_index=slide_index,
                title=title, kind=_kind(title, raw_type),
            )
        session_token = str(get("session") or "default").strip()
        session_id = stable_id("session", presentation_id, session_token)
        response_id = stable_id("response", presentation_id, session_token, participant, question_id)
        responses_by_id[response_id] = Response(
                response_id=response_id,
                session_id=session_id,
                question_id=question_id,
                # The database contract accepts a 32–128 character hex-only hash.
                participant_id=content_hash({"presentation_id": presentation_id, "session": session_token, "participant": participant}),
                value=answer,
                is_correct=_boolean(get("correct")),
                submitted_at=get("submitted") if hasattr(get("submitted"), "isoformat") else None,
            )
    responses = list(responses_by_id.values())
    if not responses:
        raise EmptyPresentationError(f"Recognized headers but no valid response rows: {Path(path).name}")
    return list(questions_by_key.values()), responses


def _parse_utc(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        token = str(value or "").strip()
        if not token:
            return None
        try:
            parsed = datetime.fromisoformat(token.replace("Z", "+00:00"))
        except ValueError:
            return None
    return parsed.replace(tzinfo=timezone.utc) if parsed.tzinfo is None else parsed.astimezone(timezone.utc)


def _parse_voters_export(
    sheet,
    header_row: int,
    headers: tuple[Any, ...],
    presentation_id: str,
    path: str | Path,
) -> tuple[list[Question], list[Response]]:
    normalized = [_key(value) for value in headers]
    date_column = normalized.index("date (utc)")
    session_column = normalized.index("session")
    voter_column = normalized.index("voter")

    questions: list[Question] = []
    by_column: dict[int, Question] = {}
    slide_by_group: dict[str, int] = {}
    for column, raw_header in enumerate(headers):
        header = str(raw_header or "").strip()
        if not header:
            continue
        answer_match = re.match(r"^(.*?):\s*Answer\s*$", header, re.I)
        if answer_match:
            title = answer_match.group(1).strip()
            group = f"answer:{_key(title)}"
            raw_type = ""
        elif "::" in header:
            prompt, criterion = header.split("::", 1)
            prompt = prompt.rstrip(":").strip()
            criterion = criterion.strip()
            if not prompt or not criterion or _key(criterion) in {"name", "emoji", "answer", "score"}:
                continue
            title = f"{prompt} — {criterion}"
            group = f"scale:{_key(prompt)}"
            raw_type = "scale"
        else:
            continue
        slide_index = slide_by_group.setdefault(group, len(slide_by_group) + 1)
        question = Question(
            stable_id("question", presentation_id, slide_index, title),
            presentation_id,
            slide_index,
            title,
            _kind(title, raw_type),
        )
        questions.append(question)
        by_column[column] = question

    if not questions:
        raise UnknownSchemaError(f"Recognized Voters sheet but no interactive columns: {Path(path).name}")

    responses: dict[str, Response] = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        session_token = str(row[session_column] or "").strip() if session_column < len(row) else ""
        voter_token = str(row[voter_column] or "").strip() if voter_column < len(row) else ""
        if not session_token or not voter_token:
            continue
        session_id = stable_id("session", presentation_id, session_token)
        participant_hash = content_hash({
            "presentation_id": presentation_id,
            "session": session_token,
            "participant": voter_token,
        })
        submitted_at = _parse_utc(row[date_column] if date_column < len(row) else None)
        for column, question in by_column.items():
            value = row[column] if column < len(row) else None
            if value in (None, ""):
                continue
            response_id = stable_id(
                "response", presentation_id, session_token, voter_token, question.question_id
            )
            responses[response_id] = Response(
                response_id,
                session_id,
                question.question_id,
                participant_hash,
                value,
                None,
                submitted_at,
            )
    if not responses:
        raise EmptyPresentationError(f"Recognized Voters sheet but no valid responses: {Path(path).name}")
    return questions, list(responses.values())


def _parse_matrix(sheet, header_row: int, participant_column: int, headers: tuple[Any, ...], presentation_id: str):
    questions: list[Question] = []
    by_column: dict[int, Question] = {}
    for column, header in enumerate(headers):
        if column == participant_column or header in (None, ""):
            continue
        title = str(header).strip()
        slide_index = len(questions) + 1
        question = Question(
            stable_id("question", presentation_id, slide_index, title), presentation_id,
            slide_index, title, _kind(title, ""),
        )
        questions.append(question)
        by_column[column] = question
    responses: dict[str, Response] = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        participant = str(row[participant_column] or "").strip() if participant_column < len(row) else ""
        if not participant:
            continue
        participant_hash = content_hash({"presentation_id": presentation_id, "session": "default", "participant": participant})
        for column, question in by_column.items():
            value = row[column] if column < len(row) else None
            if value in (None, ""):
                continue
            session_token = "default"
            response_id = stable_id("response", presentation_id, session_token, participant, question.question_id)
            responses[response_id] = Response(
                response_id, stable_id("session", presentation_id, session_token), question.question_id,
                participant_hash, value, None,
            )
    if not questions:
        raise UnknownSchemaError("Recognized matrix headers but no questions")
    if not responses:
        raise EmptyPresentationError("Recognized matrix headers but no valid participant responses")
    return questions, list(responses.values())
