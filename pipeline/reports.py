from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .metrics import (
    academic_accuracy, difficulty_band, evaluation_rate, ineffective_distractors,
    nps, participation_rate, point_biserial, public_value, reinforce,
    rolling_average, wilson_interval,
)
from .models import Presentation, Question, QuestionKind, Response, Session
from .privacy import privacy_safe_corpus, valid_response_count


PUBLIC_SHEETS = ("Resumo", "Encontros", "Questões agregadas", "Avaliações", "Assuntos", "Dicionário")
PRIVATE_SHEETS = ("Sessões", "Questões", "Respostas", "Avaliações", "Feedback", "Auditoria")
NAVY = "102A43"
TEAL = "0F766E"
LIGHT_TEAL = "CCFBF1"
WHITE = "FFFFFF"


def _style(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = TEAL
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 26
    thin = Side(style="thin", color="D9E2EC")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0FDFA")
    for column in range(1, sheet.max_column + 1):
        width = min(55, max(12, max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1)) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width


def _replace_sheet(workbook: Workbook, name: str, rows: Iterable[Iterable[object]]):
    sheet = workbook.create_sheet(name)
    for row in rows:
        sheet.append(list(row))
    _style(sheet)
    return sheet


def write_report(
    path: str | Path,
    presentations: list[Presentation],
    sessions: list[Session],
    questions: list[Question],
    responses: list[Response],
    *,
    public: bool,
) -> Path:
    if public:
        presentations, sessions, questions, responses = privacy_safe_corpus(
            presentations, sessions, questions, responses
        )
    workbook = Workbook()
    workbook.remove(workbook.active)
    by_question: dict[str, list[Response]] = defaultdict(list)
    for response in responses:
        by_question[response.question_id].append(response)
    if public:
        total_participants = sum(session.participants for session in sessions)
        academic = [response for response in responses if response.is_correct is not None]
        correct = sum(response.is_correct is True for response in academic)
        evaluation_ids = {q.question_id for q in questions if q.kind in {QuestionKind.EVALUATION, QuestionKind.NPS}}
        evaluation_respondents = {r.participant_id for r in responses if r.question_id in evaluation_ids}
        academic_respondents = {r.participant_id for r in academic}
        capacity = sum(s.participants * s.interactive_slides for s in sessions)
        valid_responses = valid_response_count(responses, questions)
        summary_sheet = _replace_sheet(workbook, "Resumo", [
            ("Indicador", "Valor"),
            ("Encontros", len(sessions)),
            ("Participantes (soma por sessão)", public_value(total_participants, total_participants)),
            ("Participação", public_value(valid_responses / capacity if capacity else None, total_participants)),
            ("Taxa de avaliação", public_value(evaluation_rate(len(evaluation_respondents), total_participants), len(evaluation_respondents))),
            ("Acurácia acadêmica", public_value(academic_accuracy(correct, len(academic)), len(academic_respondents))),
        ])
        for row in range(4, 7):
            summary_sheet.cell(row, 2).number_format = "0.0%"
        session_rows = []
        session_accuracies = []
        for session in sorted(sessions, key=lambda item: item.session_date):
            sample = [r for r in responses if r.session_id == session.session_id]
            session_valid_responses = valid_response_count(sample, questions)
            session_academic = [r for r in sample if r.is_correct is not None]
            accuracy = academic_accuracy(sum(r.is_correct is True for r in session_academic), len(session_academic))
            visible = session.participants >= 5
            session_accuracies.append(accuracy if visible else None)
            session_rows.append([session.session_date, session.presentation_id, session.participants if visible else "<5", session.interactive_slides, session_valid_responses if visible else None, participation_rate(session_valid_responses, session.participants, session.interactive_slides) if visible else None, accuracy if visible else None, None, session.complete])
        for row, moving in zip(session_rows, rolling_average(session_accuracies, 8)):
            row[7] = moving
        session_sheet = _replace_sheet(workbook, "Encontros", [("Data", "Apresentação", "Participantes", "Slides interativos", "Respostas válidas", "Participação", "Acurácia", "Média móvel (8)", "Completo")] + session_rows)
        for row in range(2, session_sheet.max_row + 1):
            session_sheet.cell(row, 1).number_format = "dd/mm/yyyy"
            for column in (6, 7, 8):
                session_sheet.cell(row, column).number_format = "0.0%"
        total_score = Counter()
        for response in academic:
            total_score[response.participant_id] += int(response.is_correct is True)
        question_rows = [("Questão", "Assunto", "N", "Corretas", "Acurácia", "Acurácia pública", "Dificuldade", "IC95% mín", "IC95% máx", "Correlação ponto-bisserial", "Distratores ineficazes", "Reforçar")]
        for question in questions:
            sample = [r for r in by_question[question.question_id] if r.is_correct is not None]
            n = len(sample)
            privacy_n = len({r.participant_id for r in sample})
            correct = sum(r.is_correct is True for r in sample)
            accuracy = academic_accuracy(correct, n)
            low, high = wilson_interval(correct, n)
            answer_counts = Counter(str(r.value) for r in sample)
            correct_choice = question.choices[question.correct_indices[0]] if question.correct_indices and question.correct_indices[0] < len(question.choices) else ""
            pb = point_biserial([bool(r.is_correct) for r in sample], [total_score[r.participant_id] for r in sample])
            distractors = ineffective_distractors(answer_counts, correct_choice)
            question_rows.append((question.title, question.topic, n if privacy_n >= 5 else "<5", correct if privacy_n >= 5 else None, accuracy if privacy_n >= 5 else None, public_value(accuracy, privacy_n), difficulty_band(accuracy) if accuracy is not None and privacy_n >= 5 else None, public_value(low, privacy_n), public_value(high, privacy_n), public_value(pb, privacy_n), ", ".join(distractors) if privacy_n >= 5 else None, reinforce(accuracy, n) if privacy_n >= 5 else False))
        question_sheet = _replace_sheet(workbook, "Questões agregadas", question_rows)
        for row in range(2, question_sheet.max_row + 1):
            for column in (5, 6, 8, 9):
                question_sheet.cell(row, column).number_format = "0.0%"
            question_sheet.cell(row, 10).number_format = "0.000"
        evaluation_questions = [q for q in questions if q.kind in {QuestionKind.EVALUATION, QuestionKind.NPS}]
        _replace_sheet(workbook, "Avaliações", [("Critério", "N", "Média pública")] + [
            (q.title, len(by_question[q.question_id]) if len({r.participant_id for r in by_question[q.question_id]}) >= 5 else "<5", public_value(nps([r.value for r in by_question[q.question_id] if isinstance(r.value, (int, float))]) if q.kind == QuestionKind.NPS else _numeric_mean(by_question[q.question_id]), len({r.participant_id for r in by_question[q.question_id]})))
            for q in evaluation_questions
        ])
        topic_counts = Counter(q.topic or q.taxonomy or "Não classificado" for q in questions if q.kind == QuestionKind.ACADEMIC)
        _replace_sheet(workbook, "Assuntos", [("Assunto", "Questões")] + [(topic, count if count >= 5 else "<5") for topic, count in sorted(topic_counts.items())])
        _replace_sheet(workbook, "Dicionário", [
            ("Campo", "Definição"),
            ("Acurácia acadêmica", "Respostas corretas / respostas válidas; exclui perfil e avaliação."),
            ("Participação", "Respostas válidas / (participantes × slides interativos)."),
            ("Privacidade", "Sessões e recortes com n < 5 são excluídos dos detalhes e dos totais públicos."),
            ("IC95%", "Intervalo de confiança binomial de Wilson a 95%."),
        ])
    else:
        private_session_sheet = _replace_sheet(workbook, "Sessões", [("ID", "Apresentação", "Data", "Participantes", "Slides interativos", "Completo")] + [
            (s.session_id, s.presentation_id, s.session_date, s.participants, s.interactive_slides, s.complete) for s in sessions
        ])
        for row in range(2, private_session_sheet.max_row + 1):
            private_session_sheet.cell(row, 3).number_format = "dd/mm/yyyy"
        _replace_sheet(workbook, "Questões", [("ID", "Apresentação", "Slide", "Questão", "Tipo", "Papel analítico", "Assunto primário", "Subassunto", "Tarefa cognitiva", "Bloom", "Dificuldade prevista", "Confiança", "Justificativa IA", "Status IA", "Versão taxonomia", "Revisar", "Revisado por", "Revisado em", "Notas da revisão")] + [
            (q.question_id, q.presentation_id, q.slide_index, q.title, q.kind.value, q.analysis_role, q.topic, q.subtopic, q.cognitive_task, q.bloom, q.predicted_difficulty, q.ai_confidence, q.ai_rationale, q.ai_status, q.taxonomy_version, q.needs_review, q.reviewed_by, q.reviewed_at.isoformat() if q.reviewed_at else None, q.review_notes) for q in questions
        ])
        _replace_sheet(workbook, "Respostas", [("ID", "Sessão", "Questão", "Participante pseudônimo", "Valor", "Correta")] + [
            (r.response_id, r.session_id, r.question_id, r.participant_id, r.value, r.is_correct) for r in responses
        ])
        _replace_sheet(workbook, "Avaliações", [("Questão", "N", "Média")] + [
            (q.title, len(by_question[q.question_id]), _numeric_mean(by_question[q.question_id])) for q in questions if q.kind in {QuestionKind.EVALUATION, QuestionKind.NPS}
        ])
        _replace_sheet(workbook, "Feedback", [("Sessão", "Questão", "Feedback")])
        _replace_sheet(workbook, "Auditoria", [("Entidade", "Quantidade"), ("Apresentações", len(presentations)), ("Sessões", len(sessions)), ("Questões", len(questions)), ("Respostas", len(responses))])
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _numeric_mean(responses: list[Response]) -> float | None:
    values = [float(r.value) for r in responses if isinstance(r.value, (int, float))]
    return sum(values) / len(values) if values else None
